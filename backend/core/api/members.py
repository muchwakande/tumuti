import io
from typing import List, Optional
from ninja import Router, File
from ninja.files import UploadedFile
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..models import FamilyMember
from ..schemas import (
    FamilyMemberCreate,
    FamilyMemberUpdate,
    FamilyMemberOut,
    FamilyMemberTreeOut,
    MessageOut,
    BulkUploadResult,
    BulkUploadError,
)
from ..auth import AuthBearer

router = Router(auth=AuthBearer())


def _apply_spouse(member: FamilyMember, new_spouse_id: Optional[int]) -> None:
    """Synchronise a bidirectional spouse relationship."""
    old_spouse = member.spouse

    # Sever the old relationship on both sides
    if old_spouse and old_spouse.pk != new_spouse_id:
        old_spouse.spouse = None
        old_spouse.save(update_fields=['spouse'])

    if new_spouse_id is None:
        member.spouse = None
        member.save(update_fields=['spouse'])
        return

    new_spouse = get_object_or_404(FamilyMember, id=new_spouse_id)

    if new_spouse.pk == member.pk:
        raise ValueError("A member cannot be their own spouse")

    # If new_spouse is already married to someone else, sever that too
    if new_spouse.spouse and new_spouse.spouse_id != member.pk:
        new_spouse.spouse.spouse = None
        new_spouse.spouse.save(update_fields=['spouse'])

    member.spouse = new_spouse
    member.save(update_fields=['spouse'])
    new_spouse.spouse = member
    new_spouse.save(update_fields=['spouse'])


def build_tree(member: FamilyMember) -> FamilyMemberTreeOut:
    """Recursively build a tree node for a member."""
    children_nodes = [build_tree(child) for child in member.children.all()]
    return FamilyMemberTreeOut(
        id=member.id,
        name=member.name,
        email=member.email,
        phone=member.phone,
        is_host=member.is_host,
        is_active=member.is_active,
        parent_id=member.parent_id,
        spouse_id=member.spouse_id,
        spouse_name=member.spouse_name,
        children=children_nodes,
    )


@router.get("/template/")
def download_template(request):
    """Download an Excel template for bulk member upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = ["name", "phone", "email", "is_host", "parent_name", "notes"]
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample rows
    samples = [
        ["Jane Doe", "+254 700 000 001", "jane@example.com", "TRUE", "", ""],
        ["John Doe", "+254 700 000 002", "", "FALSE", "Jane Doe", "Son of Jane"],
    ]
    for row_data in samples:
        ws.append(row_data)

    # Notes sheet explaining the columns
    ws_notes = wb.create_sheet("Instructions")
    instructions = [
        ("Column", "Required", "Description"),
        ("name", "Yes", "Full name of the family member"),
        ("phone", "Yes", "Phone number"),
        ("email", "No", "Email address (must be unique if provided)"),
        ("is_host", "No", "TRUE or FALSE — whether the member has opted in as a host (default: FALSE)"),
        ("parent_name", "No", "Exact name of the parent member (must already exist or appear earlier in this sheet)"),
        ("notes", "No", "Any additional notes"),
    ]
    for row_data in instructions:
        ws_notes.append(row_data)
    ws_notes.column_dimensions['A'].width = 15
    ws_notes.column_dimensions['B'].width = 10
    ws_notes.column_dimensions['C'].width = 70

    # Column widths on main sheet
    col_widths = [25, 20, 30, 10, 25, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="members_template.xlsx"'
    return response


@router.post("/bulk-upload/", response={200: BulkUploadResult, 400: MessageOut})
def bulk_upload_members(request, file: UploadedFile = File(...)):
    """Upload an Excel file to create multiple family members at once."""
    if not file.name.endswith(('.xlsx', '.xls')):
        return 400, {"message": "Only .xlsx or .xls files are accepted"}

    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
    except Exception:
        return 400, {"message": "Could not read the file. Make sure it is a valid Excel file."}

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return 400, {"message": "The file has no data rows (only a header or is empty)"}

    header_row = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    required_cols = {"name", "phone"}
    if not required_cols.issubset(set(header_row)):
        return 400, {"message": f"Missing required columns: {required_cols - set(header_row)}"}

    col = {name: idx for idx, name in enumerate(header_row)}

    def cell(row, column_name):
        idx = col.get(column_name)
        if idx is None:
            return None
        val = row[idx]
        return str(val).strip() if val is not None else ""

    # First pass: collect names created this session so parent_name can reference them
    name_to_id: dict[str, int] = {
        m.name: m.id for m in FamilyMember.objects.only("id", "name")
    }

    created_count = 0
    skipped_count = 0
    errors: list[BulkUploadError] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue  # skip fully blank rows

        name = cell(row, "name")
        phone = cell(row, "phone")
        email = cell(row, "email") or None
        raw_is_host = cell(row, "is_host") or ""
        parent_name = cell(row, "parent_name") or ""
        notes = cell(row, "notes") or ""

        if not name:
            errors.append(BulkUploadError(row=row_idx, message="name is required"))
            skipped_count += 1
            continue
        if not phone:
            errors.append(BulkUploadError(row=row_idx, message="phone is required"))
            skipped_count += 1
            continue

        is_host = raw_is_host.upper() in ("TRUE", "YES", "1")

        # Resolve parent
        parent = None
        if parent_name:
            parent_id = name_to_id.get(parent_name)
            if parent_id is None:
                errors.append(BulkUploadError(row=row_idx, message=f"parent '{parent_name}' not found"))
                skipped_count += 1
                continue
            try:
                parent = FamilyMember.objects.get(id=parent_id)
            except FamilyMember.DoesNotExist:
                errors.append(BulkUploadError(row=row_idx, message=f"parent '{parent_name}' not found"))
                skipped_count += 1
                continue

        member = FamilyMember.objects.create(
            name=name,
            phone=phone,
            email=email if email else None,
            is_host=is_host,
            parent=parent,
            notes=notes,
        )
        name_to_id[member.name] = member.id
        created_count += 1

    return 200, BulkUploadResult(
        created_count=created_count,
        skipped_count=skipped_count,
        errors=errors,
    )


@router.get("/", response=List[FamilyMemberOut])
def list_members(
    request,
    is_host: Optional[bool] = None,
    is_active: Optional[bool] = None,
):
    """List all family members with optional filters."""
    queryset = FamilyMember.objects.select_related('spouse').all()
    if is_host is not None:
        queryset = queryset.filter(is_host=is_host)
    if is_active is not None:
        queryset = queryset.filter(is_active=is_active)
    return queryset


@router.get("/tree/", response=List[FamilyMemberTreeOut])
def get_family_tree(request):
    """Return tree structure of family members (roots only, children nested)."""
    roots = FamilyMember.objects.filter(parent__isnull=True).select_related('spouse').prefetch_related(
        'children__spouse',
        'children__children__spouse',
        'children__children__children__spouse',
        'children__children__children__children__spouse',
    )
    return [build_tree(member) for member in roots]


@router.get("/{member_id}/", response={200: FamilyMemberOut, 404: MessageOut})
def get_member(request, member_id: int):
    """Get a specific family member by ID."""
    member = get_object_or_404(FamilyMember.objects.select_related('spouse'), id=member_id)
    return member


@router.post("/", response={201: FamilyMemberOut, 400: MessageOut})
def create_member(request, payload: FamilyMemberCreate):
    """Create a new family member."""
    if payload.spouse_id and payload.spouse_id == payload.parent_id:
        return 400, {"message": "Parent and spouse cannot be the same person"}

    parent = None
    if payload.parent_id:
        parent = get_object_or_404(FamilyMember, id=payload.parent_id)

    member = FamilyMember.objects.create(
        name=payload.name,
        email=payload.email,
        phone=payload.phone,
        is_host=payload.is_host,
        parent=parent,
        is_active=payload.is_active,
        notes=payload.notes,
    )

    if payload.spouse_id:
        try:
            _apply_spouse(member, payload.spouse_id)
        except ValueError as e:
            member.delete()
            return 400, {"message": str(e)}

    member.refresh_from_db()
    return 201, member


@router.patch("/{member_id}/", response={200: FamilyMemberOut, 404: MessageOut, 400: MessageOut})
def update_member(request, member_id: int, payload: FamilyMemberUpdate):
    """Update a family member."""
    member = get_object_or_404(FamilyMember.objects.select_related('spouse'), id=member_id)

    data = payload.dict(exclude_unset=True)

    # Handle parent change
    if 'parent_id' in data:
        if data['parent_id'] is not None:
            if data['parent_id'] == member_id:
                return 400, {"message": "A member cannot be their own parent"}
            parent = get_object_or_404(FamilyMember, id=data['parent_id'])
            member.parent = parent
        else:
            member.parent = None
        del data['parent_id']

    # Handle spouse change
    if 'clear_spouse' in data:
        del data['clear_spouse']
    if payload.clear_spouse:
        _apply_spouse(member, None)
    elif 'spouse_id' in data:
        new_spouse_id = data.pop('spouse_id')
        if new_spouse_id == member_id:
            return 400, {"message": "A member cannot be their own spouse"}
        try:
            _apply_spouse(member, new_spouse_id)
        except ValueError as e:
            return 400, {"message": str(e)}

    for attr, value in data.items():
        setattr(member, attr, value)
    member.save()
    member.refresh_from_db()
    return member


@router.delete("/{member_id}/", response={200: MessageOut, 404: MessageOut})
def delete_member(request, member_id: int):
    """Delete a family member."""
    member = get_object_or_404(FamilyMember, id=member_id)
    # Sever spouse link before deletion so the other side is also cleared
    if member.spouse_id:
        _apply_spouse(member, None)
    member.delete()
    return {"message": "Family member deleted successfully"}


@router.post("/{member_id}/set-spouse/", response={200: FamilyMemberOut, 400: MessageOut, 404: MessageOut})
def set_spouse(request, member_id: int, spouse_id: int):
    """Set a bidirectional spouse relationship."""
    member = get_object_or_404(FamilyMember.objects.select_related('spouse'), id=member_id)
    if spouse_id == member_id:
        return 400, {"message": "A member cannot be their own spouse"}
    try:
        _apply_spouse(member, spouse_id)
    except ValueError as e:
        return 400, {"message": str(e)}
    member.refresh_from_db()
    return member


@router.delete("/{member_id}/set-spouse/", response={200: MessageOut, 404: MessageOut})
def unset_spouse(request, member_id: int):
    """Remove a spouse relationship (both sides)."""
    member = get_object_or_404(FamilyMember.objects.select_related('spouse'), id=member_id)
    _apply_spouse(member, None)
    return {"message": "Spouse relationship removed"}
